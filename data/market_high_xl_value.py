import time
import openpyxl as op


class LTP(object):
    '''
        in this class i'm taking "ltp rows value" from exal file, sheet name is Market Highlights (Value)
    '''

    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    ltp_values = []
    for i in range(3, 385):
        LTP = sh1[f'C{i}'].value
        ltp_values.append(LTP)


class HIGH(object):
    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    high_values = []
    for i in range(3, 385):
        HIGH = sh1[f'D{i}'].value
        high_values.append(HIGH)


class LOW(object):
    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    low_values = []
    for i in range(3, 385):
        LOW = sh1[f'E{i}'].value
        low_values.append(LOW)


class CLOSEP(object):
    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    closep_values = []
    for i in range(3, 385):
        CLOSEP = sh1[f'F{i}'].value
        closep_values.append(CLOSEP)


class YCP(object):
    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    ycp_values = []
    for i in range(3, 385):
        YCP = sh1[f'G{i}'].value
        ycp_values.append(YCP)


class CHANGE(object):
    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    change_values = []
    for i in range(3, 385):
        CHANGE = sh1[f'H{i}'].value
        change_values.append(CHANGE)


class TRADE(object):
    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    trade_values = []
    for i in range(3, 385):
        TRADE = sh1[f'I{i}'].value
        trade_values.append(TRADE)


class VALUES(object):
    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    value_row_values = []
    for i in range(3, 385):
        VALUES = sh1[f'J{i}'].value
        value_row_values.append(VALUES)


class VOLUME(object):
    exl = op.load_workbook(r'/home/abdullah/Desktop/dhaka_stok/files/Book1.xlsx')
    sh1 = exl['Market Highlights (Value)']
    volume_row_values = []
    for i in range(3, 385):
        VOLUME = sh1[f'K{i}'].value
        volume_row_values.append(VOLUME)
